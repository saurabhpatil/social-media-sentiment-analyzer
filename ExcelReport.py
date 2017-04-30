__author__ = 'saurabh'

import openpyxl, json, subprocess
import ETLOperations

#   This function opens the excel workbook to modify its contents
#   It extracts data from the mongoDB collection based on the keyword entered and
#   enters the data to the excel file by matching specific fields of the aggregates
def displayExcel(searchTerm):
    workbook = openpyxl.load_workbook("ExcelReport.xlsx")
    worksheet = workbook.get_sheet_by_name('Sheet1')

    for row in range(1,1000):
        for col in range(1,12):
            worksheet.cell(row=row, column=col).value = " "

    worksheet.cell(row=1, column=1).value = 'S.No'
    # worksheet.cell(row=1, column=2).value = 'Tweet'
    worksheet.cell(row=1, column=3).value = 'Location'
    worksheet.cell(row=1, column=4).value = 'Time_Zone'
    worksheet.cell(row=1, column=5).value = 'Positive'
    worksheet.cell(row=1, column=6).value = 'Neutral'
    worksheet.cell(row=1, column=7).value = 'Negative'
    worksheet.cell(row=1, column=8).value = 'Created At'
    worksheet.cell(row=1, column=9).value = 'Followers'
    worksheet.cell(row=1, column=10).value = 'User ID'
    worksheet.cell(row=1, column=11).value = 'Product'


    dataCollection = ETLOperations.DBConnection(searchTerm)
    dataDocuments = dataCollection.find()
    wb_list=[]
    columnDict = {'positive':5,'neutral':6,'negative':7}

    for row, document in enumerate(dataDocuments,start=2):
        try:
            worksheet.cell(row=row, column=1).value = row
            worksheet.cell(row=row, column=3).value = document['user']['location']
            # worksheet.cell(row=row, column=2).value = document['text']
            worksheet.cell(row=row, column=4).value = document['user']['time_zone']
            worksheet.cell(row=row, column=columnDict[document['sentiment']['type']]).value = abs(round(float(document['sentiment']['score'])*10))
            worksheet.cell(row=row, column=8).value = document['created_at']
            worksheet.cell(row=row, column=9).value = document['user']['followers_count']
            worksheet.cell(row=row, column=10).value = document['id_str']

            wb_list.append(document)
        except:
            worksheet.cell(row=row, column=columnDict['neutral']).value = 1
            pass

    # command to launch excel sheet from within the python program
    workbook.save('ExcelReport.xlsx')
    subprocess.getoutput("TwitterAnalysisReport.xlsx")